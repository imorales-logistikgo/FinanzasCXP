from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from PendientesEnviar.models import FacturasxProveedor, RelacionFacturaProveedorxPartidas, RelacionConceptoxProyecto, View_ReporteFacturasCXP
from usersadmon.models import Proveedor
from django.template.loader import render_to_string
import json, datetime
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment


@login_required

def ReporteFacturas(request):
	if request.user.roles == 'Proveedor':
		Facturas = View_ReporteFacturasCXP.objects.filter(IDProveedor = request.user.IDTransportista).exclude(Status = 'DEPURADO')
		# listFacturas = FacturasToList(Facturas)
		ContadorPendientes, ContadorPagadas, ContadorAbonadas, ContadorCanceladas = GetContadores()
		Proveedores = Proveedor.objects.all()
		return render(request, 'ReporteFacturas.html', {'Facturas': Facturas, 'Proveedores': Proveedores, 'ContadorPagadas': ContadorPagadas, 'ContadorAbonadas': ContadorAbonadas, 'ContadorCanceladas': ContadorCanceladas, 'Rol': request.user.roles})
	else:
		Facturas = View_ReporteFacturasCXP.objects.exclude(Status__in= ('CANCELADA','DEPURADO'))
		# listFacturas = FacturasToList(Facturas)
		ContadorPendientes, ContadorPagadas, ContadorAbonadas, ContadorCanceladas = GetContadores()
		Proveedores = Proveedor.objects.all()
		return render(request, 'ReporteFacturas.html', {'Facturas': Facturas, 'Proveedores': Proveedores, 'ContadorPagadas': ContadorPagadas, 'ContadorAbonadas': ContadorAbonadas, 'ContadorCanceladas': ContadorCanceladas, 'ContadorPendientes': ContadorPendientes, 'Rol': request.user.roles})



def FacturasToList(Facturas):
	listFacturas = list()
	for Fact in Facturas:
		Factura = {}
		conFacturaxPartidas = RelacionFacturaProveedorxPartidas.objects.filter(IDFacturaxProveedor = Fact.IDFactura).select_related('IDPendienteEnviar')
		Factura['Folio'] = Fact.Folio
		Factura['Proveedor'] = Fact.NombreCortoProveedor
		Factura['FechaFactura'] = Fact.FechaFactura
		Factura["Subtotal"] = Fact.Subtotal
		Factura["IVA"] = Fact.IVA
		Factura["Retencion"] = Fact.Retencion
		Factura["Status"] = Fact.Status
		Factura['Total'] = Fact.Total
		Factura['Viajes'] = ''
		for PENDIENTE in conFacturaxPartidas:
			Factura['Viajes'] += PENDIENTE.IDPendienteEnviar.Folio + ", "
		Factura['Viajes'] = Factura['Viajes'][:-2]
		Factura["ComentarioBaja"] = Fact.ComentarioBaja
		listFacturas.append(Factura)
	return listFacturas



def GetContadores():
	AllFacturas = list(View_ReporteFacturasCXP.objects.values("Status").all())
	ContadorPendientes = len(list(filter(lambda x: x["Status"] == "PENDIENTE", AllFacturas)))
	ContadorPagadas = len(list(filter(lambda x: x["Status"] == "PAGADA", AllFacturas)))
	ContadorAbonadas = len(list(filter(lambda x: x["Status"] == "ABONADA", AllFacturas)))
	ContadorCanceladas = len(list(filter(lambda x: x["Status"] == "CANCELADA", AllFacturas)))
	return ContadorPendientes, ContadorPagadas, ContadorAbonadas, ContadorCanceladas


def GetFacturasByFilters(request):
	Proveedores = json.loads(request.GET["Proveedor"])
	Moneda = json.loads(request.GET["Moneda"])
	Status = json.loads(request.GET["Status"])
	if "Year" in request.GET:
		arrMonth = json.loads(request.GET["arrMonth"])
		Year = request.GET["Year"]
		Facturas = View_ReporteFacturasCXP.objects.filter(FechaFactura__month__in = arrMonth, FechaFactura__year = Year, Status__in = Status)#.exclude(Status = 'DEPURADO')
	else:
		Facturas = View_ReporteFacturasCXP.objects.filter(FechaFactura__range = [datetime.datetime.strptime(request.GET["FechaFacturaDesde"],'%m/%d/%Y'), datetime.datetime.strptime(request.GET["FechaFacturaHasta"],'%m/%d/%Y')], Status__in = Status)#.exclude(Status = 'DEPURADO')
	if Proveedores:
		Facturas = Facturas.filter(NombreCortoProveedor__in = Proveedores)
	if Moneda:
		Facturas = Facturas.filter(Moneda__in = Moneda)
	# if Status:
	# 	Facturas = Facturas.filter(Status__in = Status).exclude(Status = 'DEPURADO')
	# listFacturas = FacturasToList(Facturas)
	htmlRes = render_to_string('TablaReporteFacturas.html', {'Facturas':Facturas}, request = request,)
	return JsonResponse({'htmlRes' : htmlRes})


def GetReporteTotales(request, **kwargs):
	StatusIN = kwargs.get('Status', None), kwargs.get('Status2', None), "ABONADA"
	Moneda = kwargs.get('Moneda', None)
	FechaCorte = kwargs.get('FechaCorte', None)
	print(FechaCorte)
	Mes = datetime.date.today()
	Facturas = FacturasxProveedor.objects.values('IDProveedor').distinct()
	wb = Workbook()
	ws = wb.active
	ws['A1'] = 'Proveedor'
	ws['B1'] = 'Total Vencido'
	ws['C1'] = 'Total por Vencer'
	ws['D1'] = 'Total'
	ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
	ws['B1'].font = Font(bold=True, size=12, color="FFFFFF")
	ws['C1'].font = Font(bold=True, size=12, color="FFFFFF")
	ws['D1'].font = Font(bold=True, size=12, color="FFFFFF")
	ws['A1'].fill = PatternFill(bgColor="0C46B7", fill_type = "solid")
	ws['B1'].fill = PatternFill(bgColor="0C46B7", fill_type = "solid")
	ws['C1'].fill = PatternFill(bgColor="0C46B7", fill_type = "solid")
	ws['D1'].fill = PatternFill(bgColor="0C46B7", fill_type = "solid")
	ws.column_dimensions['A'].width = 30
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 22
	ws.column_dimensions['D'].width = 20
	ws['A1'].alignment = Alignment(horizontal='center')
	ws['B1'].alignment = Alignment(horizontal='center')
	ws['C1'].alignment = Alignment(horizontal='center')
	ws['D1'].alignment = Alignment(horizontal='center')
	cont=2
	for Factura in Facturas:
		NombreProv = Proveedor.objects.get(IDTransportista = Factura['IDProveedor'])
		Total = 0
		TotalVencido = 0
		TotalPorVencer = 0
		for TotalesFacturas in FacturasxProveedor.objects.filter(IDProveedor = Factura['IDProveedor'], Status__in= StatusIN, FechaFactura__lte = FechaCorte):
			if Moneda == "MXN":
				Total = Total + TotalesFacturas.Saldo if TotalesFacturas.Moneda == 'MXN' else Total + (TotalesFacturas.Saldo*TotalesFacturas.TipoCambio) if TotalesFacturas.Moneda == 'USD' else Total + TotalesFacturas.Saldo
			if TotalesFacturas.FechaVencimiento.strftime('%Y-%m-%d') <= FechaCorte and Moneda == "MXN":
				TotalVencido = TotalVencido + TotalesFacturas.Saldo if TotalesFacturas.Moneda == 'MXN' else TotalVencido + (TotalesFacturas.Saldo*TotalesFacturas.TipoCambio) if TotalesFacturas.Moneda == 'USD' else TotalVencido + TotalesFacturas.Saldo
			if TotalesFacturas.FechaVencimiento.strftime('%Y-%m-%d') > FechaCorte and Moneda == "MXN":
				TotalPorVencer = TotalPorVencer + TotalesFacturas.Saldo if TotalesFacturas.Moneda == 'MXN' else TotalPorVencer + (TotalesFacturas.Saldo*TotalesFacturas.TipoCambio) if TotalesFacturas.Moneda == 'USD' else TotalPorVencer + TotalesFacturas.Saldo
		ws.cell(row=cont,column=1).value = NombreProv.RazonSocial
		ws.cell(row=cont,column=2).value = round(TotalVencido,2)
		ws.cell(row=cont,column=3).value = round(TotalPorVencer,2)
		ws.cell(row=cont,column=4).value = round(Total,2)
		cont = cont + 1
	nombre_archivo ="ReporteFacturas.xlsx"
	response = HttpResponse(content_type="application/ms-excel")
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response
