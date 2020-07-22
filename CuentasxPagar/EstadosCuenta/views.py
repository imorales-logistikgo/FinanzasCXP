from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from bkg_viajes.models import Bro_Viajes, Bro_ServiciosxViaje, Servicios, Clientes, Bro_RepartosxViaje
from XD_Viajes.models import XD_Viajes, XD_AccesoriosxViajes, RepartosxViaje
from PendientesEnviar.models import PartidaProveedor, RelacionFacturaProveedorxPartidas, FacturasxProveedor, PendientesEnviar, RelacionConceptoxProyecto, Ext_PendienteEnviar_Costo, View_PendientesEnviarCxP, Ext_PendienteEnviar_Precio
from EstadosCuenta.models import TempSerie, View_FacturasxProveedor, PagosxProveedor, PagosxFacturas, RelacionPagosFacturasxProveedor, HistorialReajusteProveedor
from usersadmon.models import Proveedor, AdmonUsuarios, AdmonCorreosxTransportista
from users.models import User
from django.core.mail import send_mail, EmailMessage
from django.template.loader import render_to_string
from decimal import Decimal
from django.db.models import Q
import json, datetime, math
from django.contrib.auth.decorators import login_required
from django.db import transaction, DatabaseError
import json
from string import digits
from django.conf import settings
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment


@login_required
def EstadosdeCuenta(request):
	if request.user.roles == 'Proveedor':
		 return render(request, '404.html')
	else:
		result = View_FacturasxProveedor.objects.filter(Status__in = ("PENDIENTE", "ABONADA", 'APROBADA'), FechaFactura__month = datetime.datetime.now().month, FechaFactura__year = datetime.datetime.now().year)
		#result = View_FacturasxProveedor.objects.filter(Q(Status = "PENDIENTE") | Q(Status = "ABONADA") & Q(FechaFactura__month = datetime.datetime.now().month))
		ListaFacturas = FacturasToList(result)
		#Folios = list()
		#for Factura in ListaFacturas:
		#	FoliosPago = ""
		#	for Pago in RelacionPagosFacturasxProveedor.objects.filter(IDFactura = Factura["IDFactura"]).select_related('IDPago').select_related('IDFactura'):
		#		if Pago.IDPago.Status != "CANCELADA":
		#			print(str(Factura['IDFactura']) in str(Pago.IDFactura.IDFactura))
		#			FoliosPago += Pago.IDPago.Folio + ", "
		#	FoliosPago = FoliosPago[:-2]
		#	Folios.append(FoliosPago) 'Folios':Folios,
		ContadoresPendientes, ContadoresAbonadas, ContadoresPagadas, ContadoresCanceladas, ContadoresAprobada = GetContadores()
		Proveedores = Proveedor.objects.all()
		return render(request, 'EstadosdeCuenta.html', {'Facturas': ListaFacturas, 'Proveedores': Proveedores, 'ContadoresPendientes': ContadoresPendientes, 'ContadoresAbonadas': ContadoresAbonadas, 'ContadoresPagadas': ContadoresPagadas, 'ContadoresCanceladas': ContadoresCanceladas, 'ContadoresAprobada':ContadoresAprobada ,'Rol': request.user.username})


def GetDataReajuste(request):
	IDFac = request.GET["IDFactura"]
	IDPECXP = RelacionFacturaProveedorxPartidas.objects.get(IDFacturaxProveedor = IDFac)
	projectType = PendientesEnviar.objects.get(IDPendienteEnviar = str(IDPECXP.IDPendienteEnviar))
	GetConcepto = RelacionConceptoxProyecto.objects.get(IDPendienteEnviar = IDPECXP.IDPendienteEnviar)
	if projectType.Proyecto == 'BKG':
		GetBkgData = Bro_Viajes.objects.get(IDBro_Viaje = GetConcepto.IDConcepto)
		DataBKG = list()
		data = {}
		data["IDViaje"] = GetBkgData.IDBro_Viaje
		data["Folio"] = GetBkgData.Folio
		data["CostoViaje"] = GetBkgData.CostoViaje
		data["CostoRecoleccion"] = GetBkgData.CostoRecoleccion
		data["CostoAccesorios"] = GetBkgData.CostoServicios
		data["CostoRepartos"] = GetBkgData.CostoTotalRepartos
		data["CostoSubtotal"] = GetBkgData.CostoSubtotal
		data["CostoIVA"] = GetBkgData.CostoIVA
		data["CostoRetencion"] = GetBkgData.CostoRetencion
		data["CostoTotal"] = GetBkgData.CostoTotal
		data["Proyecto"] = 'BKG'
		data["TipoViaje"] = GetBkgData.Tipo
		DataBKG.append(data)
	elif projectType.Proyecto == 'XD':
		GetXdData = XD_Viajes.objects.get(XD_IDViaje = GetConcepto.IDConcepto)
		DataBKG = list()
		data = {}
		data["IDViaje"] = GetXdData.XD_IDViaje
		data["Folio"] = GetXdData.Folio
		data["CostoViaje"] = GetXdData.Costo
		data["CostoAccesorios"] = GetXdData.CostoAccesorios
		data["CostoRepartos"] = GetXdData.CostoRepartos
		data["CostoSubtotal"] = GetXdData.CostoSubtotal
		data["CostoIVA"] = GetXdData.CostoIVA
		data["CostoRetencion"] = GetXdData.CostoRetencion
		data["CostoTotal"] = GetXdData.CostoTotal
		data["Proyecto"] = 'XD'
		data["TipoViaje"] = GetXdData.TipoViaje
		DataBKG.append(data)
	return JsonResponse({'DataBKG':DataBKG})


def GetAccesoriosxViaje(request):
	IDViaje_ = request.GET["IDViaje"]
	GetDataAccesoriosBKG = Bro_ServiciosxViaje.objects.filter(IDBro_Viaje = IDViaje_)
	GetDataAccesoriosXD = XD_AccesoriosxViajes.objects.filter(XD_IDViaje = IDViaje_)
	GetnameAccesorio = Servicios.objects.all()
	NewData = list()
	costoA = 0;
	dataFromProject = GetnameAccesorio if (request.GET["Proyecto"] == 'BKG')  else GetDataAccesoriosXD if (request.GET["Proyecto"] == 'XD') else ""
	for Accesorios in dataFromProject :
		dataAccesorios = {}
		if request.GET["Proyecto"] == 'BKG':
			for eachAccesorio in GetDataAccesoriosBKG:
				costoA = eachAccesorio.Costo if (Accesorios.IDservicio == eachAccesorio.IDBro_Servicio) else 0
		dataAccesorios['NombreAccesorio'] = Accesorios.Nombre if(request.GET["Proyecto"] == 'BKG') else Accesorios.Descripcion if (request.GET["Proyecto"] == 'XD') else ""
		dataAccesorios['CostoAccesorio'] = costoA if(request.GET["Proyecto"] == 'BKG') else Accesorios.Costo if (request.GET["Proyecto"] == 'XD') else ""
		dataAccesorios['IsAplicaRetencion'] = Accesorios.IsAplicaRetencion if(request.GET["Proyecto"] == 'BKG') else ""
		dataAccesorios['IsAplicaIVA'] = Accesorios.IsAplicaiva if(request.GET["Proyecto"] == 'BKG') else ""
		#dataAccesorios['TipoViaje'] = Accesorios.TipoViaje if(request.GET["Proyecto"] == 'BKG') else Accesorios.TipoViaje if(request.GET["Proyecto"] == 'XD')  else ""
		NewData.append(dataAccesorios)
	return JsonResponse({"NewData":NewData})


def GetRepartosxViaje(request):
	IDViaje_ = request.GET['IDViaje']
	GetDataRepartosBKG = Bro_RepartosxViaje.objects.filter(IDBro_Viaje = IDViaje_)
	GetDataRepartosXD = RepartosxViaje.objects.filter(XD_IDViaje = IDViaje_)
	NewDataR = list()
	whichProyect =  GetDataRepartosBKG if (request.GET["Proyecto"] == 'BKG')  else GetDataRepartosXD if (request.GET["Proyecto"] == 'XD') else ""
	for Repartos in whichProyect:
		dataRepartos = {}
		dataRepartos["deliveries"] = 0 if (request.GET["Proyecto"] == 'BKG') else Repartos.Numero if (request.GET["Proyecto"] == 'XD') else ""
		try:
			GetDestinoReparto = Clientes.objects.get(IDCliente = Repartos.IDCliente) if (request.GET["Proyecto"] == 'BKG') else ""
			dataRepartos["ciudadDestino"] = GetDestinoReparto.Estado if (request.GET["Proyecto"] == 'BKG') else Repartos.Estado if (request.GET["Proyecto"] == 'XD') else ""
		except Clientes.DoesNotExist:
			dataRepartos["ciudadDestino"] = ""
		dataRepartos["costo"] = Repartos.CostoReparto if (request.GET["Proyecto"] == 'BKG') else Repartos.Costo if (request.GET["Proyecto"] == 'XD') else ""
		NewDataR.append(dataRepartos)
	return JsonResponse({"NewDataR":NewDataR})



def saveReajuste(request):
	jParams = json.loads(request.body.decode('utf-8'))
	try:
		idPendienteEnviarReajuste = RelacionFacturaProveedorxPartidas.objects.get(IDFacturaxProveedor = jParams["IDFactura"])
		FacturaReajuste = FacturasxProveedor.objects.get(IDFactura = jParams["IDFactura"])
		with transaction.atomic(using='users'):
			newFacturaReajuste = HistorialReajusteProveedor()
			newFacturaReajuste.IDPendienteEnviar = str(idPendienteEnviarReajuste.IDPendienteEnviar)
			newFacturaReajuste.IDFacturaxProveedor = jParams["IDFactura"]
			newFacturaReajuste.CostoSubtotalAnterior = FacturaReajuste.Subtotal
			newFacturaReajuste.CostoIVAAnterior =FacturaReajuste.IVA
			newFacturaReajuste.CostoRetencionAnterior =FacturaReajuste.Retencion
			newFacturaReajuste.CostoTotalAnterior = FacturaReajuste.Total
			newFacturaReajuste.NuevoCosto =jParams["Costo"]
			newFacturaReajuste.NuevoCostoRecoleccion = jParams["CostoRecoleccion"]
			newFacturaReajuste.NuevoCostoRepartos = jParams["CostoRepartos"]
			newFacturaReajuste.NuevoCostoAccesorios = jParams["CostoAccesorios"]
			newFacturaReajuste.NuevoCostoSubtotal = jParams["Subtotal"]
			newFacturaReajuste.NuevoCostoIVA = jParams["IVA"]
			newFacturaReajuste.NuevoCostoRetencion = jParams["Retencion"]
			newFacturaReajuste.NuevoCostoTotal = jParams["Total"]
			newFacturaReajuste.FechaAlta = datetime.datetime.now()
			newFacturaReajuste.IDUsuarioAlta = request.user.idusuario
			newFacturaReajuste.Motivo = jParams["Motivo"]
			newFacturaReajuste.save()
			FacturaReajuste.Subtotal = jParams["Subtotal"]
			FacturaReajuste.IVA = jParams["IVA"]
			FacturaReajuste.Retencion = jParams["Retencion"]
			FacturaReajuste.Total = jParams["Total"]
			FacturaReajuste.Saldo = jParams["Total"]
			FacturaReajuste.save()
			DataFacturaByID = list()
			NewDataFactura = FacturasxProveedor.objects.get(IDFactura = jParams["IDFactura"])
			data = {}
			data["newSubtotal"] = NewDataFactura.Subtotal
			data["newIVA"] = NewDataFactura.IVA
			data["newRetencion"] = NewDataFactura.Retencion
			data["newTotal"] = NewDataFactura.Total
			DataFacturaByID.append(data)
			return JsonResponse({'DataFacturaByID':DataFacturaByID})
	except:
		transaction.rollback(using = 'users')
		return HttpResponse(status=400)





def GetContadores():
	AllFacturas = list(View_FacturasxProveedor.objects.values('Status').all())
	ContadoresPendientes = len(list(filter(lambda x: x["Status"] == "PENDIENTE", AllFacturas)))
	ContadoresAbonadas = len(list(filter(lambda x: x["Status"] == "ABONADA", AllFacturas)))
	ContadoresPagadas = len(list(filter(lambda x: x["Status"] == "PAGADA", AllFacturas)))
	ContadoresCanceladas = len(list(filter(lambda x: x["Status"] == "CANCELADA", AllFacturas)))
	ContadoresAprobada = len(list(filter(lambda x: x["Status"] == "APROBADA", AllFacturas)))
	return ContadoresPendientes, ContadoresAbonadas, ContadoresPagadas, ContadoresCanceladas, ContadoresAprobada



def GetFacturasByFilters(request):
	Proveedores = json.loads(request.GET["Proveedor"])
	Status = json.loads(request.GET["Status"])
	Moneda = json.loads(request.GET["Moneda"])
	if "Year" in request.GET:
		arrMonth = json.loads(request.GET["arrMonth"])
		Year = request.GET["Year"]
		Facturas = View_FacturasxProveedor.objects.filter(FechaFactura__month__in = arrMonth, FechaFactura__year = Year, Status__in = Status)#.exclude(Status = 'DEPURADO')
	else:
		Facturas = View_FacturasxProveedor.objects.filter(FechaFactura__range = [datetime.datetime.strptime(request.GET["FechaFacturaDesde"],'%m/%d/%Y'), datetime.datetime.strptime(request.GET["FechaFacturaHasta"],'%m/%d/%Y')], Status__in = Status)#.exclude(Status = 'DEPURADO')
	# if Status:
	# 	Facturas = Facturas.filter(Status__in = Status)
	if Proveedores:
		Facturas = Facturas.filter(Proveedor__in = Proveedores)
	if Moneda:
		Facturas = Facturas.filter(Moneda__in = Moneda)
	ListaFacturas = FacturasToList(Facturas)
	#Folios = list()
	#for Factura in ListaFacturas:
	#	FoliosPago= ""
	#	for Pago in RelacionPagosFacturasxProveedor.objects.filter(IDFactura = Factura["IDFactura"]).select_related('IDPago'):
	#		FoliosPago += Pago.IDPago.Folio + ", "
	#	FoliosPago = FoliosPago[:-2]
	#	Folios.append(FoliosPago)
	htmlRes = render_to_string('TablaEstadosCuenta.html', {'Facturas': ListaFacturas,}, request = request,)
	return JsonResponse({'htmlRes' : htmlRes})



def FacturasToList(Facturas):
	ListaFacturas = list()
	for Factura in Facturas:
		NuevaFactura = {}
		NuevaFactura["IDFactura"] = Factura.IDFactura
		NuevaFactura["Folio"] = Factura.Folio
		NuevaFactura["Proveedor"] = Factura.Proveedor
		NuevaFactura["FechaFactura"] = Factura.FechaFactura
		NuevaFactura["Subtotal"] = Factura.Subtotal
		NuevaFactura["IVA"] = Factura.IVA
		NuevaFactura["Retencion"] = Factura.Retencion
		NuevaFactura["Total"] = Factura.Total
		NuevaFactura["Saldo"] = Factura.Saldo
		NuevaFactura["FechaFactura"] = Factura.FechaFactura
		NuevaFactura["Moneda"] = Factura.Moneda
		NuevaFactura["Status"] = Factura.Status
		NuevaFactura["RutaXML"] = Factura.RutaXML
		NuevaFactura["IsAutorizada"] = Factura.IsAutorizada
		NuevaFactura["IDProveedor"] = Factura.IDProveedor
		NuevaFactura["TotalXML"] = Factura.TotalXML
		NuevaFactura["Pago"] = Factura.FolioPago #if (Factura.StatusPago != 'CANCELADA') else ""
		ListaFacturas.append(NuevaFactura)
	return ListaFacturas



def CancelarFactura(request):
	IDFactura = json.loads(request.body.decode('utf-8'))["IDFactura"]
	Motivo = json.loads(request.body.decode('utf-8'))["MotivoEliminacion"]
	conRelacionFacturaProveedorxPartidas = RelacionFacturaProveedorxPartidas.objects.filter(IDFacturaxProveedor = IDFactura)
	if conRelacionFacturaProveedorxPartidas:
		conRelacionFacturaProveedorxPartidas[0].IDFacturaxProveedor.Status = 'CANCELADA'
		conRelacionFacturaProveedorxPartidas[0].IDFacturaxProveedor.IDUsuarioBaja = AdmonUsuarios.objects.get(idusuario = request.user.idusuario)
		conRelacionFacturaProveedorxPartidas[0].IDFacturaxProveedor.ComentarioBaja = Motivo
		conRelacionFacturaProveedorxPartidas[0].IDFacturaxProveedor.save()
		for Partida in conRelacionFacturaProveedorxPartidas:
			Partida.IDPartida.IsActiva = False
			Partida.IDPartida.FechaBaja = datetime.datetime.now()
			Ext_Costo = Ext_PendienteEnviar_Costo.objects.get(IDPendienteEnviar = Partida.IDPendienteEnviar)
			Ext_Costo.IsFacturaProveedor = False
			Ext_Costo.save()
			Partida.IDPartida.save()
	return HttpResponse("")




def GetDetallesFactura(request):
	ListaViajes = list()
	IDFactura = request.GET["IDFactura"]
	conRelacionFacturaProveedorxPartidas = RelacionFacturaProveedorxPartidas.objects.filter(IDFacturaxProveedor = IDFactura)
	if conRelacionFacturaProveedorxPartidas:
		for Partida in conRelacionFacturaProveedorxPartidas:
			Viaje = {}
			Pending = RelacionConceptoxProyecto.objects.get(IDPendienteEnviar = Partida.IDPendienteEnviar)
			Viaje["Folio"] = Pending.IDPendienteEnviar.Folio
			Viaje["FechaDescarga"] = Pending.IDPendienteEnviar.FechaDescarga
			Ext_Costo = Ext_PendienteEnviar_Costo.objects.get(IDPendienteEnviar = Pending.IDPendienteEnviar)
			Viaje["Subtotal"] = Ext_Costo.CostoSubtotal
			Viaje["IVA"] = Ext_Costo.CostoIVA
			Viaje["Retencion"] = Ext_Costo.CostoRetencion
			Viaje["Total"] = Ext_Costo.CostoTotal
			ListaViajes.append(Viaje)
	htmlRes = render_to_string('TablaDetallesFactura.html', {'Pendientes':ListaViajes}, request = request,)
	return JsonResponse({'htmlRes' : htmlRes})



def SavePagoxProveedor(request):
	jParams = json.loads(request.body.decode('utf-8'))
	newPago = PagosxProveedor()
	newPago.FechaAlta = datetime.datetime.now()
	newPago.Folio = jParams["Folio"]
	newPago.Total = jParams["Total"]
	newPago.FechaPago = datetime.datetime.strptime(jParams["FechaPago"],'%Y/%m/%d')
	if "RutaComprobante" in jParams:
		newPago.RutaComprobante = jParams["RutaComprobante"]
	newPago.Comentarios = jParams["Comentarios"]
	newPago.TipoCambio = jParams["TipoCambio"]
	NombrePro = Proveedor.objects.filter(IDTransportista = jParams["Proveedor"]).get()
	newPago.NombreCortoProveedor = NombrePro.NombreComercial
	newPago.IDUsuarioAlta = AdmonUsuarios.objects.get(idusuario = request.user.idusuario)
	newPago.IDProveedor = jParams["IDProveedor"]
	newPago.save()
	return HttpResponse(newPago.IDPago)



def truncate(number, digits) -> Decimal:
    stepper = 10.0 ** digits
    return math.trunc(stepper * number) / stepper


def SavePagoxFactura(request):
	jParams = json.loads(request.body.decode('utf-8'))
	for Pago in jParams["arrPagos"]:
		newPagoxFactura = PagosxFacturas()
		newPagoxFactura.Total = Pago["Total"]
		newPagoxFactura.FechaAlta = datetime.datetime.now()
		newPagoxFactura.save()
		newRelacionPagoxFactura = RelacionPagosFacturasxProveedor()
		newRelacionPagoxFactura.IDPago = PagosxProveedor.objects.get(IDPago = jParams["IDPago"])
		newRelacionPagoxFactura.IDPagoxFactura = newPagoxFactura
		Factura = FacturasxProveedor.objects.get(IDFactura = Pago["IDFactura"])
		Factura.Saldo -= Decimal(float(Pago["Total"])/float(jParams["TipoCambio"]) if (Factura.Moneda == 'USD') else Pago["Total"])
		newRelacionPagoxFactura.IDFactura = FacturasxProveedor.objects.get(IDFactura = Pago["IDFactura"])
		if truncate(float(Factura.Saldo), 2) == 0:
			Factura.Status = "PAGADA"
		else:
			Factura.Status = "ABONADA"
		Factura.save()
		newRelacionPagoxFactura.save()
	# try:
	#  	MsjCorreo = EnviarCorreoProveedor(IDPagoEmail = jParams["IDPago"])
	# except Exception as e:
	# 	pass
	return HttpResponse('')



def ValidarFactura(request):
	IDFactura = json.loads(request.body.decode('utf-8'))["IDFactura"]
	Factura = FacturasxProveedor.objects.get(IDFactura = IDFactura)
	if Factura:
		Factura.IsAutorizada = True
		Factura.Status = 'APROBADA'
		Factura.save()
	Status = Factura.Status
	return JsonResponse({"Status":Status})



def CheckFolioDuplicado(request):
	IsDuplicated = PagosxProveedor.objects.filter(Folio = request.GET["Folio"]).exclude(Status = "CANCELADA").exists()
	return JsonResponse({'IsDuplicated' : IsDuplicated})


def EnviarCorreoProveedor(IDPagoEmail):
	DatosPagoProveedor = PagosxProveedor.objects.get(IDPago = IDPagoEmail)
	try:
		CorreoProveedor = list(AdmonCorreosxTransportista.objects.filter(IDTransportista = DatosPagoProveedor.IDProveedor, IsEnviarCorreo = 1).values('Correo'))
		if CorreoProveedor != []:
			SendEmail = list()
			for new in CorreoProveedor:
				SendEmail.append(new["Correo"], )
			RS = Proveedor.objects.get(IDTransportista = DatosPagoProveedor.IDProveedor)
			context={
				'nombre': RS.RazonSocial,
				'folio' : DatosPagoProveedor.Folio,
				'fechapago' : DatosPagoProveedor.FechaPago,
				'total' : DatosPagoProveedor.Total
			}
			template_name='email.html'
			html_content=render_to_string("CorreoProveedor.html", context)
			subject='Notificación de pago'
			from_email= settings.EMAIL_HOST_USER
			to= SendEmail
			cc=[settings.EMAIL_HOST_USER]
			# reply_to=['jfraga@logisti-k.com.mx']

			msg = EmailMessage(subject, html_content, from_email, to, cc=cc)
			msg.content_subtype = "html"  # Main content is now text/html
			msg.send()
			return "Success"
		else:
			return "Error"
	except Exception as e:
		print(e)
		return "Error"



def GetDetallesPago(request):
	IDFactura = request.GET["IDFactura"]
	FacturasxPago = RelacionPagosFacturasxProveedor.objects.filter(IDFactura = IDFactura).select_related('IDPago').select_related('IDPagoxFactura')
	Facturas = list()
	for FacturaxPago in FacturasxPago:
		Pago = {}
		Pago["FolioPago"] = FacturaxPago.IDPago.Folio
		Pago["FechaPago"] = FacturaxPago.IDPago.FechaPago
		Pago["Total"] = FacturaxPago.IDPagoxFactura.Total
		Pago["Status"] = "ACTIVO" if (FacturaxPago.IDPago.Status != 'CANCELADA') else FacturaxPago.IDPago.Status
		Facturas.append(Pago)
	htmlRes = render_to_string('TablaDetallesPago.html', {'Facturas':Facturas}, request = request,)
	return JsonResponse({'htmlRes' : htmlRes})


def FixIDProveedor(request):
	Facturas = RelacionFacturaProveedorxPartidas.objects.all()
	for Factura in Facturas:
		IDProveedor = RelacionConceptoxProyecto.objects.filter(IDPendienteEnviar = Factura.IDPendienteEnviar.IDPendienteEnviar)[0].IDProveedor
		Fact = FacturasxProveedor.objects.get(IDFactura = Factura.IDFacturaxProveedor.IDFactura)
		Fact.IDProveedor = IDProveedor
		Fact.save()
	return HttpResponse('Done')

def pruebaexcel(request):
	Facturas = FacturasxProveedor.objects.values('IDProveedor','NombreCortoProveedor').distinct()

		# print(datetime.datetime.now().strftime('%Y-%m-%d') > Factura.FechaVencimiento.strftime('%Y-%m-%d'))
	wb = Workbook()
	ws = wb.active
	ws['A1'] = 'Proveedor'
	ws['B1'] = 'Total Vencido'
	ws['C1'] = 'Total por Vencer'
	ws['D1'] = 'Total'
	ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
	ws['B1'].font = Font(bold=True, size=12, color="FFFFFF")
	ws['C1'].font = Font(bold=True, size=12, color="FFFFFF")
	ws['D1'].font = Font(bold=True, size=12, color="FFFFFF")
	ws['A1'].fill = PatternFill(bgColor="0C46B7", fill_type = "solid")
	ws['B1'].fill = PatternFill(bgColor="0C46B7", fill_type = "solid")
	ws['C1'].fill = PatternFill(bgColor="0C46B7", fill_type = "solid")
	ws['D1'].fill = PatternFill(bgColor="0C46B7", fill_type = "solid")
	ws.column_dimensions['A'].width = 30
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 22
	ws.column_dimensions['D'].width = 20
	ws['A1'].alignment = Alignment(horizontal='center')
	ws['B1'].alignment = Alignment(horizontal='center')
	ws['C1'].alignment = Alignment(horizontal='center')
	ws['D1'].alignment = Alignment(horizontal='center')
	cont=2
	for Factura in Facturas:
		Total = 0
		TotalVencido = 0
		TotalPorVencer = 0
		for TotalesFacturas in FacturasxProveedor.objects.filter(IDProveedor = Factura['IDProveedor'], Status__in = ('PENDIENTE','APROVADA')):
			Total = Total + TotalesFacturas.Total
			if datetime.datetime.now().strftime('%Y-%m-%d') > TotalesFacturas.FechaVencimiento.strftime('%Y-%m-%d'):
				TotalVencido = TotalVencido + TotalesFacturas.Total
			if datetime.datetime.now().strftime('%Y-%m-%d') < TotalesFacturas.FechaVencimiento.strftime('%Y-%m-%d'):
				TotalPorVencer = TotalPorVencer + TotalesFacturas.Total
		ws.cell(row=cont,column=1).value = Factura['NombreCortoProveedor']
		ws.cell(row=cont,column=2).value = '$' + str(round(TotalVencido,2))
		ws.cell(row=cont,column=3).value = '$' + str(round(TotalPorVencer,2))
		ws.cell(row=cont,column=4).value = '$' + str(round(Total,2))
		cont = cont + 1
	nombre_archivo ="ReporteFacturas.xlsx"
	response = HttpResponse(content_type="application/ms-excel")
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response
# def InsertSerieProveedor(request):
# 	Serie = list(FacturasxProveedor.objects.values('IDProveedor').distinct())
# 	for each in Serie:
# 		try:
# 			ser = FacturasxProveedor.objects.filter(IDProveedor = each["IDProveedor"])[:1].get()
# 			remove_digits = str.maketrans('', '', digits)
# 			res = ser.Folio.translate(remove_digits)
# 			saveP = Proveedor.objects.get(IDTransportista=each["IDProveedor"])
# 			saveP.Serie = res
# 			saveP.save()
# 		except Exception as e:
# 			print(e)

def leerExcel(reques):
	archivo_excel = pd.read_excel('static/json/UUIDErrorMAriel.xlsx')
	# values = archivo_excel['Factura']
	try:
		for i in archivo_excel.index:
			a = FacturasxProveedor.objects.filter(IDFactura = archivo_excel['IDFactura'][i]).exclude(Status = 'CANCELADA').get()
			a.UUID = archivo_excel['Correcto'][i]
			a.save()
			print(a.IDFactura)
			# if a.UUID is None:
				# a.UUID = archivo_excel['UUID'][i]
				# a.save()
				# print(a.Folio)
	except Exception as e:
		print(archivo_excel['IDFactura'][i])
		print(e)

	# 	z = {}
	# 	z['Folio'] = a.Folio
	# 	z['FechaDescarga'] = a.FechaDescarga
	# 	z['IsFacturaProveedor'] = a.IsFacturaProveedor
	# 	b.append(z)
	# 	print(a.IsFacturaProveedor)
	# print(b)

		# separador = ";"
		# print(archivo_excel['CxP'][i])
		# if archivo_excel['CxP'][i] != 'nan':
		# 	NewData = archivo_excel['CxP'][i].split(separador)
		# 	for correo in NewData:
		# 		try:
		# 			InsertCorreo = AdmonCorreosxTransportista()
		# 			InsertCorreo.IDTransportista = Proveedor.objects.get(IDTransportista = archivo_excel['IDTransportista'][i])
		# 			InsertCorreo.Correo = correo.replace(" ", "")
		# 			InsertCorreo.IsEnviarCorreo = 0
		# 			InsertCorreo.save()
		# 		except Exception as e:
		# 			print(e)
		# else:
		# 	print(archivo_excel['IDTransportista'][i])

# 		try:
# 			with transaction.atomic(using='users'):
# 				p = PendientesEnviar()
# 				p.Folio = archivo_excel['Folio'][i]
# 				p.NombreCortoCliente = archivo_excel['Cliente']s[i]
# 				p.NombreCortoProveedor = archivo_excel['Transportista'][i]
# 				p.FechaDescarga = archivo_excel['FechaDescarga'][i]
# 				p.Moneda = archivo_excel['Moneda'][i]
# 				p.Status = archivo_excel['statusproceso'][i]
# 				p.IsEvidenciaFisica = archivo_excel['IsEvidenciasFisicas'][i]
# 				p.IsEvidenciaDigital = archivo_excel['IsEvidenciasDigitales'][i]
# 				p.Proyecto = 'BKG'
# 				p.TipoConcepto = 'VIAJE'
# 				p.IsCrontrolDesk = archivo_excel['IsCapturaControlDesk'][i]
# 				p.save()
# 				r = RelacionConceptoxProyecto()
# 				r.IDConcepto = archivo_excel['IDConcepto'][i]
# 				r.IDCliente = archivo_excel['IDCliente'][i]
# 				r.IDProveedor = archivo_excel['IDTransportista'][i]
# 				r.IDPendienteEnviar = PendientesEnviar.objects.get(IDPendienteEnviar = p.IDPendienteEnviar)
# 				r.save()
# 				ex = Ext_PendienteEnviar_Costo()
# 				ex.CostoSubtotal = Decimal(archivo_excel['CostoSubtotal'][i].item())
# 				ex.CostoIVA = Decimal(archivo_excel['CostoIVA'][i].item())
# 				ex.CostoRetencion = Decimal(archivo_excel['CostoRetencion'][i].item())
# 				ex.CostoTotal = Decimal(archivo_excel['CostoTotal'][i].item())
# 				ex.IsFacturaProveedor = True if (archivo_excel['Factura'][i] != '') else False
# 				ex.IDPendienteEnviar = PendientesEnviar.objects.get(IDPendienteEnviar = p.IDPendienteEnviar)
# 				ex.save()
# 				# exp = Ext_PendienteEnviar_Precio()
# 				# exp.PrecioSubtotal = Decimal(archivo_excel['PrecioSubtotal'][i].item())
# 				# exp.PrecioIVA = Decimal(archivo_excel['PrecioIVA'][i].item())
# 				# exp.PrecioRetencion = Decimal(archivo_excel['PrecioRetencion'][i].item())
# 				# exp.PrecioTotal = Decimal(archivo_excel['PrecioTotal'][i].item())
# 				# exp.IsFacturaCliente = False
# 				# exp.IDPendienteEnviar = PendientesEnviar.objects.get(IDPendienteEnviar = p.IDPendienteEnviar)
# 				# exp.save()
# 				if FacturasxProveedor.objects.filter(Folio = archivo_excel['Factura'][i]).exclude(Status = "CANCELADA").exists():
# 					idF = FacturasxProveedor.objects.filter(Folio = archivo_excel['Factura'][i]).exclude(Status = "CANCELADA").get()
# 					idF.Subtotal = (Decimal(archivo_excel['CostoSubtotal'][i].item()) + idF.Subtotal)
# 					idF.IVA = (Decimal(archivo_excel['CostoIVA'][i].item()) + idF.IVA)
# 					idF.Retencion = (Decimal(archivo_excel['CostoRetencion'][i].item()) + idF.Retencion)
# 					idF.Total = ((Decimal(archivo_excel['CostoSubtotal'][i].item()) + idF.Subtotal)+(Decimal(archivo_excel['CostoIVA'][i].item()) + idF.IVA)) - (Decimal(archivo_excel['CostoRetencion'][i].item()) + idF.Retencion)
# 					idF.Saldo = ((Decimal(archivo_excel['CostoSubtotal'][i].item()) + idF.Subtotal)+(Decimal(archivo_excel['CostoIVA'][i].item()) + idF.IVA)) - (Decimal(archivo_excel['CostoRetencion'][i].item()) + idF.Retencion)
# 					idF.save()
# 					pr = PartidaProveedor()
# 					pr.FechaAlta = '2020-01-01'
# 					pr.Subtotal = Decimal(archivo_excel['CostoSubtotal'][i].item())
# 					pr.IVA = Decimal(archivo_excel['CostoIVA'][i].item())
# 					pr.Retencion = Decimal(archivo_excel['CostoRetencion'][i].item())
# 					pr.Total = Decimal(archivo_excel['CostoTotal'][i].item())
# 					pr.IsActiva = True
# 					pr.save()
# 					RF = RelacionFacturaProveedorxPartidas()
# 					RF.IDFacturaxProveedor = FacturasxProveedor.objects.get(IDFactura = idF.IDFactura)
# 					RF.IDPartida = PartidaProveedor.objects.get(IDPartida = pr.IDPartida)
# 					RF.IDPendienteEnviar = PendientesEnviar.objects.get(IDPendienteEnviar = p.IDPendienteEnviar)
# 					RF.save()
# 				else:
# 					fac = FacturasxProveedor()
# 					fac.Folio = archivo_excel['Factura'][i]
# 					fac.NombreCortoProveedor = archivo_excel['Transportista'][i]
# 					fac.FechaFactura = '2020-01-01'
# 					fac.FechaRevision = '2020-01-01'
# 					fac.FechaVencimiento = '2020-01-01'
# 					fac.Moneda = archivo_excel['Moneda'][i]
# 					fac.Subtotal = Decimal(archivo_excel['CostoSubtotal'][i].item())
# 					fac.IVA = Decimal(archivo_excel['CostoIVA'][i].item())
# 					fac.Retencion = Decimal(archivo_excel['CostoRetencion'][i].item())
# 					fac.Total = Decimal(archivo_excel['CostoTotal'][i].item())
# 					fac.Saldo = Decimal(archivo_excel['CostoTotal'][i].item())
# 					fac.IsAutorizada = False
# 					fac.RutaXML = ''
# 					fac.RutaPDF = ''
# 					fac.TipoCambio = 1
# 					fac.Comentarios = ''
# 					fac.TotalConvertido = 0
# 					fac.Status = 'PENDIENTE'
# 					fac.IDUsuraioAlta = 152
# 					fac.IDProveedor = archivo_excel['IDTransportista'][i]
# 					fac.save()
# 					pr = PartidaProveedor()
# 					pr.FechaAlta = '2020-01-01'
# 					pr.Subtotal = Decimal(archivo_excel['CostoSubtotal'][i].item())
# 					pr.IVA = Decimal(archivo_excel['CostoIVA'][i].item())
# 					pr.Retencion = Decimal(archivo_excel['CostoRetencion'][i].item())
# 					pr.Total = Decimal(archivo_excel['CostoTotal'][i].item())
# 					pr.IsActiva = True
# 					pr.save()
# 					RF = RelacionFacturaProveedorxPartidas()
# 					RF.IDFacturaxProveedor = FacturasxProveedor.objects.get(IDFactura = fac.IDFactura)
# 					RF.IDPartida = PartidaProveedor.objects.get(IDPartida = pr.IDPartida)
# 					RF.IDPendienteEnviar = PendientesEnviar.objects.get(IDPendienteEnviar = p.IDPendienteEnviar)
# 					RF.save()
# 		except Exception as e:
# 			transaction.rollback(using='users')
# 			print(e)
	#for para leer cada row del excel
	# f = list()
	# for Excel in archivo_excel.index:
	# 	try:
	# 		a = FacturasxProveedor.objects.filter(Folio = archivo_excel['FACTURA'][Excel]).get()
	# 		if a.Moneda==archivo_excel['Moneda'][Excel]:
	# 			print('Yes')
	# 		else:
	# 			a.Moneda = archivo_excel['Moneda'][Excel]
	# 			a.save()
	# 			print('save')
	# 	except Exception as e:
	# 		print(e)
	# print(f)
	#
	# f = list()
	# for data in values:
	# 	try:
	# 		a = PendientesEnviar.objects.filter(Folio = data)[:1].get()
	# 		print("Yes")
	# 		# a.FechaFactura = '2020-01-01'
	# 		# a.save()
	# 	except:
	# 		f.append(data, )
	# print(f)
